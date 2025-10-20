"""
Doimiy Excel fayl - har doim yangilanadi
Yangi booking qo'shilganda avtomatik yoziladi
"""
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font
from db import get_all_bookings, get_user_by_id, get_booking_by_id
from config import PRICES
import datetime
import os
import logging

logging.basicConfig(level=logging.INFO)

EXCEL_FILE = "bookings.xlsx"  # XLSX ga o'zgartirdim (CSV emas)

# Shablon kengliklari (yangi Telefon ustuni qo'shildi)
COLUMN_WIDTHS = {
    'A': 5,   # №
    'B': 25,  # FISH
    'C': 12,  # Sana
    'D': 15,  # Vaqt
    'E': 15,  # Maydon
    'F': 15,  # Pul
    'G': 15   # Telefon
}

def load_or_create_excel():
    """Excel ni load qilish yoki yangisini yaratish (shablon bilan)"""
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Sheet1'
        # Header qo'shish (yangi Telefon ustuni bilan)
        ws['A1'] = 'Butun tarix'
        ws.merge_cells('A1:G1')  # Yangilandi: F dan G gacha
        ws.append(['№', 'FISH', 'Sana', 'Vaqt', 'Maydon', 'Pul', 'Telefon'])
        apply_styles(ws)
        wb.save(EXCEL_FILE)
    else:
        wb = openpyxl.load_workbook(EXCEL_FILE)
    return wb

def apply_styles(ws):
    """All borders, alignment va font ni qo'llash (chiroyli qilish uchun)"""
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=7):  # Yangilandi: max_col 6 dan 7 ga
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if cell.row == 1 or cell.row == 2:  # Headerlarni bold qilish
                cell.font = Font(bold=True)

    # Kengliklarni o'rnatish (katakcha uzunligi)
    for col, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width

    # Qator balandligi
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 15  # Standard Excel height

def sort_by_date(ws):
    """Sana bo'yicha tartiblash (3-qatordan boshlab)"""
    # Ma'lumotlarni list ga olish
    data = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=7, values_only=True):  # Yangilandi: max_col 6 dan 7 ga
        data.append(list(row))

    # Sana (C ustuni, index 2) bo'yicha tartiblash ('dd.mm.yyyy' stringdan datetime ga)
    def date_key(x):
        try:
            return datetime.datetime.strptime(x[2], '%d.%m.%Y')
        except:
            return datetime.datetime.min  # Agar xato bo'lsa, eng eski

    data.sort(key=date_key)

    # Eski qatorlarni o'chirish
    ws.delete_rows(3, ws.max_row - 2 if ws.max_row > 2 else 0)

    # Yangi qatorlarni qo'shish va № ni yangilash
    for idx, row_data in enumerate(data, start=1):
        ws.append([idx] + row_data[1:])

    apply_styles(ws)  # Stilni qayta qo'llash

def append_booking_to_excel(booking_id):
    """
    Yangi booking qo'shilganda Excel ga yozish va tartiblash
    """
    booking = get_booking_by_id(booking_id)
    if not booking:
        logging.error(f"O'yin ID {booking_id} topilmadi!")
        return

    user = get_user_by_id(booking.user_id)
    if not user:
        logging.error(f"Foydalanuvchi topilmadi!")
        return

    wb = load_or_create_excel()
    ws = wb['Sheet1']

    # Sana ni string formatda yozish
    date_formatted = booking.date.strftime('%d.%m.%Y')
    full_name = f"{user.name} {user.surname}"

    # Qo'shish oldin, agar allaqachon mavjud bo'lsa, qo'shmaslik
    exists = False
    for row in range(3, ws.max_row + 1):
        if (ws.cell(row, 3).value == date_formatted and
            ws.cell(row, 4).value == booking.slot and
            ws.cell(row, 5).value == booking.field and
            ws.cell(row, 2).value == full_name and
            ws.cell(row, 7).value == user.phone):  # Telefonni ham tekshirish
            exists = True
            logging.warning(f"Duplicate booking ID {booking_id} allaqachon Excel da mavjud! Qo'shilmadi.")
            break

    if not exists:
        # Yangi qator qo'shish (Telefon qo'shildi)
        new_row = [
            0,  # Vaqtinchalik № (sort da yangilanadi)
            full_name,
            date_formatted,
            booking.slot,
            booking.field,
            PRICES.get(booking.field, 0),
            user.phone  # Yangi Telefon ustuni
        ]
        ws.append(new_row)
        logging.info(f"✅ Yangi o'yin {booking_id} qo'shildi")

    sort_by_date(ws)  # Sana bo'yicha tartiblash va № ni yangilash
    wb.save(EXCEL_FILE)

    logging.info(f"✅ O'yin {booking_id} Excel ga qo'shildi (duplicate bo'lmasa)")

def delete_booking_from_excel(booking_id):
    """
    Booking ni o'chirish va qayta tartiblash
    """
    booking = get_booking_by_id(booking_id)
    if not booking:
        logging.error(f"O'yin ID {booking_id} topilmadi!")
        return

    user = get_user_by_id(booking.user_id)
    if not user:
        logging.error(f"Foydalanuvchi topilmadi!")
        return

    full_name = f"{user.name} {user.surname}"

    wb = load_or_create_excel()
    ws = wb['Sheet1']

    date_formatted = booking.date.strftime('%d.%m.%Y')

    # Qidirish va barcha mos qatorlarni o'chirish
    row = 3
    found_count = 0
    while row <= ws.max_row:
        if (ws.cell(row, 3).value == date_formatted and
            ws.cell(row, 4).value == booking.slot and
            ws.cell(row, 5).value == booking.field and
            ws.cell(row, 2).value == full_name and
            ws.cell(row, 7).value == user.phone):  # Telefonni ham tekshirish
            ws.delete_rows(row)
            found_count += 1
            logging.info(f"O'chirildi: row {row} for booking {booking_id}")
            continue
        row += 1

    if found_count == 0:
        logging.warning(f"O'yin ID {booking_id} Excel da topilmadi! (Full name: {full_name}, Phone: {user.phone})")
    else:
        logging.info(f"✅ {found_count} ta mos qator o'chirildi for booking {booking_id}")

    sort_by_date(ws)  # Qayta tartiblash va № ni yangilash
    wb.save(EXCEL_FILE)

def generate_full_excel():
    """
    To'liq Excel faylni qayta yaratish (barcha ma'lumotlar bilan)
    """
    from db import get_all_bookings, get_user_by_id

    bookings = get_all_bookings()
    if not bookings:
        logging.info("Ma'lumot yo'q")
        return "Ma'lumot yo'q"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # Header (Telefon qo'shildi)
    ws['A1'] = 'Butun tarix'
    ws.merge_cells('A1:G1')
    ws.append(['№', 'FISH', 'Sana', 'Vaqt', 'Maydon', 'Pul', 'Telefon'])

    # Ma'lumotlarni qo'shish
    for b in bookings:
        user = get_user_by_id(b.user_id)
        date_formatted = b.date.strftime('%d.%m.%Y')
        full_name = f"{user.name} {user.surname}" if user else "N/A"
        new_row = [
            0,  # Vaqtinchalik №
            full_name,
            date_formatted,
            b.slot,
            b.field,
            PRICES.get(b.field, 0),
            user.phone if user else "N/A"  # Yangi Telefon ustuni
        ]
        ws.append(new_row)

    sort_by_date(ws)  # Tartiblash
    wb.save(EXCEL_FILE)

    logging.info(f"✅ To'liq Excel yaratildi: {len(bookings)} ta o'yin")
    return f"✅ Excel yaratildi: {len(bookings)} ta o'yin"

def get_excel_file():
    """
    Excel faylni qaytarish (yuklab olish uchun, bytes)
    """
    if not os.path.exists(EXCEL_FILE):
        generate_full_excel()

    with open(EXCEL_FILE, 'rb') as f:
        return f.read()